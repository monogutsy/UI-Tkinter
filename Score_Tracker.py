import openpyxl
from openpyxl import load_workbook
from openpyxl import Workbook
import tkinter as tk
from tkinter import ttk
from tkinter import messagebox
import os

file_name="Student_Scores.xlsx"

def submit():
    student_name=entry1.get()
    score=entry2.get()
    
    if not student_name or not score:
        messagebox.showwarning("Please input name and score")
        return
    if not score.isdigit():
        messagebox.showwarning("Invalid Score")
        return
    score=int(score)
    if score < 0 or score > 100:
        messagebox.showwarning("Invalid Score")
        return
    status="Fail" if score < 75 else "Pass"
    if not os.path.exists(file_name):
        wb=Workbook()
        ws=wb.active
        ws.title = "Scores"
        ws.append(["Student Name", "Score", "Status"])  
        wb.save(file_name)

    messagebox.showinfo("Success")
    wb =load_workbook(file_name)
    ws= wb.active
    ws.append([student_name, score, status])
    wb.save(file_name)
    messagebox.showinfo("Success")
    entry1.delete(0, tk.END)
    entry2.delete(0, tk.END)
    refresh_data()
    scores = [ws.cell(row=p,column=2).value for p in range(2, ws.max_row +1) if isinstance(ws.cell(row=p, column=2).value, int)]
    if score:
        avrscore= sum(scores)/len(scores)
        avrstatus= "Fail" if avrscore < 75 else "Pass"
        ws.append(["Average Score",avrscore, avrstatus])
        





def refresh_data():
    for row in var1.get_children():
        var1.delete(row)
    if os.path.exists(file_name):
        wb = load_workbook(file_name)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] and isinstance(row[1], int): 
                var1.insert("", "end", values=row)



window = tk.Tk()
window.geometry("650x500")
window.title("Score Tracker")


frame1=tk.Frame(window)
frame1.grid(row=1, column=1, padx=110, pady=20)

label1= tk.Label(frame1, text="Student Name:",font="Elephant")
label1.grid(row=0, column=0, padx=10, pady=10)
entry1 = tk.Entry(frame1)
entry1.insert(0,"Name")
entry1.grid(row=0, column=1, padx=10, pady=10)

label2= tk.Label(frame1, text="Score:",font="Elephant")
label2.grid(row=1, column=0, padx=10, pady=10)
entry2= tk.Entry(frame1)
entry2.insert(0,"Score (0-100)")
entry2.grid(row=1, column=1, padx=10, pady=10)

button1 =tk.Button(frame1, text="Submit", command=submit)
button1.grid(row=2, column=1, padx=10, pady=10)

var1 = ttk.Treeview(window,columns=("Student Name", "Score", "Status"), show="headings")
var1.grid(row=3, column=1, padx=20, pady=20)

var1.heading("Student Name", text="Student Name")
var1.heading("Score", text="Score")
var1.heading("Status", text="Status")

refresh_data()





window.mainloop()